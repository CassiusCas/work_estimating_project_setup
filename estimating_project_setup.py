#NEXT: copy brs file template to prj folder then replace paths in program to point to new location
#NEXT: transfer user input information to gensum
#next: transfer brs first column company to gensum
#next: transfer final brs results to historical database 
#import libraries
import numpy as np
import pandas as pd
import urllib
import os
import openpyxl
#import winshell
from win32com.client import Dispatch
from shutil import copy as cp
import datetime
#from uszipcode import SearchEngine, SimpleZipcode, Zipcode


def import_csv(file):
    #import csv file
    df=pd.read_csv(file)
    return df
def get_download_path():
    #return default download path for linux or windows
        if os.name == 'nt':
            import winreg
            sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
            downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
                location = winreg.QueryValueEx(key, downloads_guid)[0]
            return location
        else:
            return os.path.join(os.path.expanduser('~'), 'downloads')


def master_program():
    #get user input project information to fill into excel documents accross the board
    print("After you type each piece of information requested. Press 'Enter' on your keyboard.")
    prj_name=input("Project Name:\n")
    prj_numb=input("Project Number:\n")
    bid_date=input("Bid Date:\n")
    location=input("Location:\n")
    t_office=input("Turner Office:\n")
    client=input("Client Name:\n")
    lead_est=input("Lead Estimator:\n")

    #get user download directory on windows computer
    download_path = get_download_path()

    #combines users default download path with standard file name
    project_setup_directory= download_path +r'\_01-estimating_project_setup'
    prj_folder = project_setup_directory+r"\\"+prj_numb+r"__-__"+prj_name
    templates_path = project_setup_directory+r"\template files"
    #create project folder if doesn't already exist
    if os.path.exists(prj_folder):
        return
    else:
        os.makedirs(prj_folder)
                   #copy template files into project folder and rename for project
    time_now = datetime.datetime.now()
    current_time =str(time_now)[0:19].replace(":","-")
    #brs_prj_file = 
    gensum_prj_file = prj_folder+ r"\\"+current_time+"__-__gensum.xlsx"
    #check if template files for project have already been copied and renamed, if yes copies another file and adds version to end 
    #incase someone runs program twice won't overwrite existing excel files
    cp(templates_path+r"\gensum_template_full.xlsx",gensum_prj_file)

    brs_temp_url=project_setup_directory +r'\brs_template.xlsx'
    print(brs_temp_url)
    #open brs template file
    wb_brs = openpyxl.load_workbook(brs_temp_url)

    for sheet in wb_brs:
        print(sheet.title)




    #####ENTER PROJECT INFORMATION INTO BRS WORKSHEET#####
    ws_p_info = wb_brs.worksheets[0]
    ws_p_info["B2"] = prj_name
    ws_p_info["B3"] = prj_numb
    ws_p_info["B4"] = bid_date
    ws_p_info["B5"] = location
    ws_p_info["B6"] = t_office
    ws_p_info["B7"] = client
    ws_p_info["B8"] = lead_est

    #save brs
    new_brs = project_setup_directory + r'\new_brs.xlsx'
    wb_brs.save(new_brs)
    #
    #
    #
    #
    #
    #

    #instead of having the user go and input the file name everytime when they download the file from building connected
    #they should always download it to their downloads folder and call it estimating_project_setup.csv
    csv_path = project_setup_directory + r'\eps.csv'
    print(csv_path)

    #read csv file into pandas to convert to excel xlsx file
    df = import_csv(csv_path)
   
    #create excel dataframe with only trade number designation and trade package
    df2 = df.loc[:,('Number','Bid Package','Bid Package Lead')]

    #drop duplicate rows based on number and bid package column
    df2.drop_duplicates(subset=['Number','Bid Package'],inplace=True)

    #test:print all rows of dataframe
    #print(df2)

    #create Sub Corespondance folder
    sb_path = project_setup_directory +r'\sub_correspondence'
    if not os.path.exists(sb_path):
        os.makedirs(sb_path)
    
    
    path_test = sb_path + r'\test'
    shell = dispatch('WScript.Shell')
    shortcut = shell.CreateShortcut(path_test)
    #shortcut.Targetpath = 

    #find max length of the longest string in trade column
    #for each line of trade length subtract from len_trade and add that many spaces. trying to line up all names
    #len_trade = df2.loc[:,'Bid Package'].map(lambda x: len(x)).max()

    #create bid package folders under sub corespondance with naming convention "number - Trade - Estimating lead 
    for x in range(len(df2)):
        #char_length_package = len(df2.iloc[x,1])
        #space_multi = (len_trade - char_length_package)*'_'
        bid_path = sb_path + '\\'+df2.iloc[x,0] +'_'+df2.iloc[x,1] + ' - ' +df2.iloc[x,2]
        bid_path_rep = bid_path.replace('=', '').replace('"','')
        print(bid_path_rep)
        os.makedirs(bid_path_rep,exist_ok=True)
     
    for x in range(len(df2)):
        #creating a standard name of excel file
        file_name = download_path+r'\_01-estimating_project_setup\eps.xlsx'

    #create excel file from pandas
    df.to_excel(file_name,sheet_name='project_setup')

    #import newly saved excel file with openpyXL for manipulation
    wb = openpyxl.load_workbook(file_name)

    #test if open py excel has properly imported excel doc
    try:
        for sheet in wb:
            print(sheet.title)
    except:
        print("Error when opening workbook with OpenPyXL")
    else:
        print("workbook successfully imported with OpenPyXl")

    for x in range(len(df2)):
        sheet_name = df2.iloc[x,0] +' - '+df2.iloc[x,1]
        sheet_name_rep = sheet_name.replace('=', '').replace('"','').replace('(','').replace(')','').replace('and','&').replace(',','')
        wb.create_sheet(sheet_name_rep)

    for sheet in wb:
        print(sheet.title)

    wb.save(file_name)


master_program()